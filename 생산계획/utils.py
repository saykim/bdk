from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os

def generate_pdf(orders, output_path):
    """생산계획을 PDF로 출력"""
    doc = SimpleDocTemplate(
        output_path,
        pagesize=landscape(A4),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )
    
    # 테이블 데이터 준비
    elements = []
    headers = ['공장', '제품', '생산오더번호', '수량', '시작시간', '종료시간', '상태']
    data = [headers]
    
    for order in orders:
        data.append([
            order.factory,
            order.product,
            order.order_no,
            f"{order.quantity:,.0f}",
            order.start_time.strftime('%Y-%m-%d %H:%M'),
            order.end_time.strftime('%Y-%m-%d %H:%M'),
            order.status
        ])
    
    # 테이블 스타일 설정
    table = Table(data)
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ])
    table.setStyle(style)
    elements.append(table)
    
    # PDF 생성
    doc.build(elements)

def generate_excel(orders, output_path):
    """생산계획을 Excel로 출력"""
    wb = Workbook()
    ws = wb.active
    ws.title = "생산계획"
    
    # 헤더 설정
    headers = ['공장', '제품', '생산오더번호', '수량', '시작시간', '종료시간', '상태']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # 데이터 입력
    for row, order in enumerate(orders, 2):
        ws.cell(row=row, column=1, value=order.factory)
        ws.cell(row=row, column=2, value=order.product)
        ws.cell(row=row, column=3, value=order.order_no)
        ws.cell(row=row, column=4, value=order.quantity)
        ws.cell(row=row, column=5, value=order.start_time.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row, column=6, value=order.end_time.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row, column=7, value=order.status)
        
        # 셀 스타일 적용
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal='center')
    
    # 컬럼 너비 자동 조정
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # 파일 저장
    wb.save(output_path)

def ensure_upload_folder():
    """업로드 폴더가 없으면 생성"""
    if not os.path.exists('static/exports'):
        os.makedirs('static/exports') 